from autoaddress20 import Autoaddress
from os import getenv
from dotenv import load_dotenv
from io import BytesIO
from openpyxl import load_workbook
from msoffcrypto import OfficeFile      # unfortunately openpyxl doesn't handle encrypted workbooks ref: https://stackoverflow.com/a/66879133
from json import loads

def result_to_excel(addressData):
    if addressData.get("reformattedAddress", 0):
        formattedAddress = addressData["reformattedAddress"]
        ws['D' + str(i)] = "1"
        ws['E' + str(i)] = formattedAddress[0]
        if formattedAddress[1]:
            ws['F' + str(i)] = formattedAddress[1]
        if formattedAddress[2]:
            ws['G' + str(i)] = formattedAddress[2]
        ws['H' + str(i)] = formattedAddress[3]
        if formattedAddress[4]:
            ws['I' + str(i)] = formattedAddress[4]
        else:
            ws['I' + str(i)] = formattedAddress[3]      # County town, only town name given
    else:
        ws['D' + str(i)] = "2"

load_dotenv()
api_key = getenv('AUTOADDRESS_APIKEY_DEV')        # Development key
#api_key = getenv('AUTOADDRESS_APIKEY_PROD')      # Production key

a = Autoaddress(api_key, api_url="https://api-bulk.autoaddress.ie/2.0/")         # change the api URL to separate from regular website usage
#a.FindAddress(getenv('TEST_ADDRESS'))
#print(a.address)

decrypted_workbook = BytesIO()
address_file = "address_list.xlsx"
with open(address_file, 'rb') as file:
    office_file = OfficeFile(file)
    office_file.load_key(password=getenv('SHEET_PASSWORD'))
    office_file.decrypt(decrypted_workbook)
wb = load_workbook(filename=decrypted_workbook)

ws = wb['Sheet1']
# Add headers
ws['D1'] = "LookupStatus"
ws['E1'] = "Address1"
ws['F1'] = "Address2"
ws['G1'] = "Address3"
ws['H1'] = "Town"
ws['I1'] = "County"

for i in range(2,2063):
    eircode = ws['C' + str(i)].value
    print(eircode)
    a.FindAddress(eircode)
    #print(a.address)
    addressData = loads(a.address)
    result_code = addressData["result"]["code"]
    result_text = addressData["result"]["text"]
    if result_code == 550:
        ws['D' + str(i)] = "0"
        ws['E' + str(i)] = "Address not found for this eircode"
    elif result_code == 100:
        print("Address found")
        result_to_excel(addressData)
    elif result_code == 600:            # ForeignAddressDetected
        a.FindAddress(eircode, query_Country="ni")      # Fetch the address again with the country code set to Northern Ireland
        addressData = loads(a.address)
        result_code = addressData["result"]["code"]
        result_text = addressData["result"]["text"]
        print(result_code, result_text)
    else:
        ws['D' + str(i)] = "3"
        ws['E' + str(i)] = (f"Unknown code returned: {result_code}, {result_text}")

try:
    wb.save("/mnt/scratch/output_address_list.xlsx")
except Exception as e:
    print(e)
